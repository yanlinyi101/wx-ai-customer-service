"""
知识库管理工具
使用方式：
  python kb_tool.py list          查看所有条目
  python kb_tool.py add           交互式添加新条目
  python kb_tool.py delete <序号>  删除指定条目（序号从1开始）
  python kb_tool.py export        导出为 Excel 表格
  python kb_tool.py import        从 Excel 表格导入
"""

import json
import sys
from pathlib import Path

KB_PATH = Path(__file__).parent / "knowledge_base.json"


def load() -> list:
    if not KB_PATH.exists():
        return []
    with open(KB_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save(kb: list) -> None:
    with open(KB_PATH, "w", encoding="utf-8") as f:
        json.dump(kb, f, ensure_ascii=False, indent=2)
    print(f"✅ 已保存，共 {len(kb)} 条")


def cmd_list():
    kb = load()
    if not kb:
        print("知识库为空")
        return
    print(f"\n共 {len(kb)} 条知识库内容：\n")
    for i, entry in enumerate(kb, 1):
        keywords = "、".join(entry.get("keywords", []))
        print(f"[{i:02d}] 问：{entry['question']}")
        print(f"      答：{entry['answer'][:40]}{'...' if len(entry['answer']) > 40 else ''}")
        print(f"      关键词：{keywords}")
        print()


def cmd_add():
    print("\n=== 添加新条目 ===")
    question = input("问题描述（如：如何申请退款？）：").strip()
    if not question:
        print("问题不能为空")
        return
    answer = input("标准答案：").strip()
    if not answer:
        print("答案不能为空")
        return
    keywords_raw = input("触发关键词（用空格分隔，如：退款 退货 七天）：").strip()
    keywords = [k for k in keywords_raw.split() if k]
    image_url = input("图片链接（无图片直接回车跳过）：").strip()

    entry = {
        "question": question,
        "answer": answer,
        "keywords": keywords,
        "image_url": image_url,
    }
    kb = load()
    kb.append(entry)
    save(kb)
    print(f"\n✅ 已添加第 {len(kb)} 条：{question}")


def cmd_delete(index_str: str):
    try:
        index = int(index_str) - 1
    except ValueError:
        print("请输入有效序号（数字）")
        return
    kb = load()
    if index < 0 or index >= len(kb):
        print(f"序号超出范围，当前共 {len(kb)} 条")
        return
    removed = kb.pop(index)
    save(kb)
    print(f"✅ 已删除：{removed['question']}")


def cmd_export():
    try:
        import openpyxl
    except ImportError:
        print("请先安装 openpyxl：pip install openpyxl")
        return

    kb = load()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "知识库"
    ws.append(["问题", "答案", "关键词（空格分隔）", "图片链接（可为空）"])

    # 设置列宽
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 50

    for entry in kb:
        ws.append([
            entry.get("question", ""),
            entry.get("answer", ""),
            " ".join(entry.get("keywords", [])),
            entry.get("image_url", ""),
        ])

    out_path = Path(__file__).parent / "knowledge_base.xlsx"
    wb.save(out_path)
    print(f"✅ 已导出到：{out_path}")


def cmd_import():
    try:
        import openpyxl
    except ImportError:
        print("请先安装 openpyxl：pip install openpyxl")
        return

    xlsx_path = Path(__file__).parent / "knowledge_base.xlsx"
    if not xlsx_path.exists():
        print(f"找不到文件：{xlsx_path}\n请先运行 export 导出模板再填写")
        return

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    kb = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        question, answer, keywords_raw = row[0], row[1], row[2]
        image_url = row[3] if len(row) > 3 else ""
        if not question or not answer:
            continue
        keywords = [k for k in str(keywords_raw or "").split() if k]
        kb.append({
            "question": str(question),
            "answer": str(answer),
            "keywords": keywords,
            "image_url": str(image_url or "").strip(),
        })

    save(kb)
    print(f"✅ 已从 Excel 导入 {len(kb)} 条")


# ── 入口 ───────────────────────────────────────
if __name__ == "__main__":
    args = sys.argv[1:]

    if not args or args[0] == "list":
        cmd_list()
    elif args[0] == "add":
        cmd_add()
    elif args[0] == "delete" and len(args) > 1:
        cmd_delete(args[1])
    elif args[0] == "export":
        cmd_export()
    elif args[0] == "import":
        cmd_import()
    else:
        print(__doc__)
